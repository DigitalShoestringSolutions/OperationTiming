from django.db.models import Q
from django.http import HttpResponse
from django.conf import settings
from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes, renderer_classes
from rest_framework.permissions import IsAuthenticatedOrReadOnly, AllowAny
from rest_framework.renderers import JSONRenderer, BrowsableAPIRenderer
from rest_framework.response import Response
import datetime
import dateutil.parser
import re
import urllib

from openpyxl import load_workbook
import xlrd
import os
import tempfile

from .models import Identifier, IdentifierType, IdentifierPattern, IdentityEntry, IdentityType
from .serializers import IdentitySerializer, IdentitySerializerFull, IdentityTypeSerializer
from .forms import UploadFileForm

@api_view(('GET',))
@renderer_classes((JSONRenderer,BrowsableAPIRenderer))
def identify(request,identifier_type,identifier):
    full = request.GET.get("full",False)
    try:
        idfier_type = IdentifierType.objects.get(tag__exact=identifier_type)
    except IdentifierType.DoesNotExist:
        return Response({"reason":"Identifier Type Not Found"},status=status.HTTP_400_BAD_REQUEST)
    try:
        idfier = Identifier.objects.get(type__exact=idfier_type,value__exact=identifier)
        identity = idfier.target
        print(f"{identifier_type}:{identifier}>Found:{identity.get_id()}")
    except Identifier.DoesNotExist:
        for pattern_obj in idfier_type.patterns.all():
            pattern = re.compile(pattern_obj.pattern)
            match = pattern.match(identifier)
            if match:
                dataset = {**pattern_obj.defaults,**match.groupdict()}
                identity = IdentityEntry.objects.create(type=pattern_obj.id_type,**dataset)
                Identifier.objects.create(type=idfier_type,value=identifier,target=identity)
                print(f"{identifier_type}:{identifier}>Created:{identity.get_id()}")

    
    if full is not False:
        serializer = IdentitySerializerFull(identity)
    else:
        serializer = IdentitySerializer(identity)
    return Response(serializer.data)

@api_view(('GET',))
@renderer_classes((JSONRenderer,BrowsableAPIRenderer))
def identify_multi(request,identifier_type,id=None):
    full_ids = []
    id_nums = []
    if id is not None:
        ids = [id]
    else:
        raw_ids = request.GET.getlist("id")
        ids = [urllib.parse.unquote(id) for id in raw_ids]

    print(f'IDs: {ids}')  
    print(f'ID Type: {identifier_type}') 

    full = request.GET.get("full",False)
    try:
        idfier_type = IdentifierType.objects.get(tag__exact=identifier_type)
    except IdentifierType.DoesNotExist:
        return Response({"reason":"Identifier Type Not Found"},status=status.HTTP_400_BAD_REQUEST)
    
    for id in ids:
        try:
            idfier = Identifier.objects.get(type__exact=idfier_type,value__exact=id)
            identity = idfier.target
            full_id = identity.get_id()
            full_ids.append(full_id)
            print(f"{identifier_type}:{id}>Found:{identity.get_id()}")
        except Identifier.DoesNotExist:
            return Response({"reason":"Identity Not Found"},status=status.HTTP_400_BAD_REQUEST)

    for id_entry in full_ids:
        *id_type,id_num = id_entry.split('@')
        id_nums.append(id_num)
    
    try:
        identities = IdentityEntry.objects.filter(auto_id__in=id_nums)
    except IdentityEntry.DoesNotExist:
        return Response({"reason":"Identity Not Found"},status=status.HTTP_400_BAD_REQUEST)
    

    serializer = IdentitySerializerFull(identities,many=True)

    return Response(serializer.data)

@api_view(('GET',))
@renderer_classes((JSONRenderer,BrowsableAPIRenderer))
def getID(request,id=None):
    if id is not None:
        ids = [id]
    else:
        raw_ids = request.GET.getlist("id")
        ids = [urllib.parse.unquote(id) for id in raw_ids]
    
    print(f'IDs: {ids}')
    id_nums = []
    for id_entry in ids:
        *id_type,id_num = id_entry.split('@')
        id_nums.append(id_num)
    
    try:
        identities = IdentityEntry.objects.filter(auto_id__in=id_nums)
    except IdentityEntry.DoesNotExist:
        return Response({"reason":"Identity Not Found"},status=status.HTTP_400_BAD_REQUEST)
        
    serializer = IdentitySerializerFull(identities,many=True)

    return Response(serializer.data)



@api_view(('GET',))
@renderer_classes((JSONRenderer,BrowsableAPIRenderer))
def listByIDType(request,id_type):
    full = request.GET.get("full",False)
    try:
        id_type = IdentityType.objects.get(tag__exact=id_type)
        qs = id_type.identities.all()
        if full is not False:
            serializer = IdentitySerializerFull(qs,many=True)
        else:
            serializer = IdentitySerializer(qs,many=True)
        return Response(serializer.data)

    except IdentityType.DoesNotExist:
        return Response({"reason":"Identity Type Not Found"},status=status.HTTP_400_BAD_REQUEST)

@api_view(('GET',))
@renderer_classes((JSONRenderer,BrowsableAPIRenderer))
def listTypes(request):
    qs = IdentityType.objects.all()
    serializer = IdentityTypeSerializer(qs,many=True)
    return Response(serializer.data)

def uploadIdentities(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            results = handle_upload(request.FILES["file"])
            print(results)
            return render(request,"upload_done.html",{"results":results})
    else:
        form = UploadFileForm()
    return render(request,"upload_form.html",{"form":form})

def handle_upload(file_obj):
    acc = {'list':[],'created':[],'error':[]}
    filename = file_obj.name

    _, file_extension = os.path.splitext(filename)
    if file_extension == ".xlsx":
        wb = load_workbook(filename = file_obj)
        ws = wb.active

        cell_range = ws['A']

        for cell in cell_range[1:]:
            if cell.value is not None:
                type_tag = ws[f'A{cell.row}'].value
                id_name = ws[f'B{cell.row}'].value
                idfier_type_tag = ws[f'C{cell.row}'].value
                idfier_value = ws[f'D{cell.row}'].value

                handle_row(type_tag,id_name,idfier_type_tag,idfier_value, acc)

    elif file_extension == ".xls":
        fd,path = tempfile.mkstemp()
        try:
            with os.fdopen(fd,'wb') as tmp:
                tmp.write(file_obj.read())
            wb = xlrd.open_workboo(path)
            ws = wb.sheet_by_index(0)
            
            for nx in range(1,ws.nrows):
                cell = ws.cell(nx,0)
                if cell.ctype != 0:
                    type_tag = ws.cell(nx,0).value
                    id_name = ws.cell(nx,1).value
                    idfier_type_tag = ws.cell(nx,2).value
                    idfier_value = ws.cell(nx,3).value

                    handle_row(type_tag,id_name,idfier_type_tag,idfier_value,acc)
        finally:
            os.remove(path)

    return acc

def handle_row(ttag,id_name,idf_ttag,idf_v,acc):
    try:
        id_t = IdentityType.objects.get(tag=ttag)
        id,id_created = IdentityEntry.objects.get_or_create(name=id_name,type=id_t)

        if id_created:
            print(f"Created ID: {id_t.title}:{id_name}")
            acc['created'].append(f"ID: {id_t.title}:{id.name}")

        idf_t = IdentifierType.objects.get(tag=idf_ttag)
        idf, idf_created = Identifier.objects.get_or_create(type=idf_t,value=idf_v,target=id)

        if idf_created:
            print(f"Created IDFR: {idf_t.title}:{idf_v}")
            acc['created'].append(f"IDENTIFIER: {idf_t.title}:{idf.value}")

        acc['list'].append(f"{idf_t.title}:{idf.value} => {id_t.title}:{id.name}")
    except Exception as e:
        print(f"Failed on {ttag}:{id_name},{idf_ttag}:{idf_v}")
        print(e)
        acc['error'].append(f"{ttag}:{id_name}<=>{idf_ttag}:{idf_v}")
